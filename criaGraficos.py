'''
Autor: Luiz Fernando Antonelli Galati
Jul/2023
'''

'''
Este código lê dados de um documento no formato .xlsx, cria, de forma automatizada, vários gráficos para facilitar a visualização
dos dados do documento e, por fim, escreve os resultados (os gráficos) em uma nova planilha do próprio documento.
'''

import openpyxl
import os
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import numpy as np
import math
import time


def criaGraficosSemestres (arquivo, nomeArquivo):
    planilha2 = arquivo["Gráficos (por semestre)"]
    categoriasPlan2 = ["Domínio de\nconceitos", "Conhecimen-\nto de áreas", "Aplicação\nprática", "Pesquisa\njurídica", "Comunicação", "Colaboração", "Ética", "Empreende-\ndorismo", "Cosmopoli-\ntanismo"]
    valoresPlan2 = []
   
    if (os.path.isdir ("Gráficos") == 0):
        os.mkdir ("Gráficos")

    i = 1
    k = planilha2.max_row + 3
    l = 0
    while (i < planilha2.max_column):
        titulo = planilha2.cell (row = 1, column = i).value
        j = 3
        while (j < 12):        
            conteudo = planilha2.cell (row = j, column = i + 1).value            
            conteudo = conteudo*100                     
            valoresPlan2.append (conteudo)            
            j = j + 1 

        if (nomeArquivo == "Dados quantitativos agregados (eletivas e clínicas).xlsx"):
            caminho_fig = "Gráficos/(Eletivas) " + titulo + ".png"
        else:
            caminho_fig = "Gráficos/(Obrigatórias) " + titulo + ".png"
            
        plt.rc ('xtick', labelsize = 6.2)
        fig, ax = plt.subplots (figsize = (7.12, 3.66))
        ax.bar (categoriasPlan2, valoresPlan2, width = 0.50)
        ax.set_ylabel ("Porcentagem da pontuação máxima", fontsize = 9.16)
        ax.set_ylim([0, 100])
        fig.suptitle (titulo)        
        plt.close ()
   
        fig.savefig (caminho_fig)    
        img = openpyxl.drawing.image.Image (caminho_fig)
        if (i % 2 == 1):
            string = str (k)
            string = "A" + string                             
            planilha2.add_image (img, string)            
        else:
            string = str (k)
            string = "M" + string                             
            planilha2.add_image (img, string)
            k = k + 21

        valoresPlan2 = []
        l = l + 1       
        i = i + 3


def criaGraficosQuesitos (arquivo, nomeArquivo):
    planilha3 = arquivo["Gráficos (por quesito)"]
    categoriasPlan3 = []
    valoresPlan3 = []

    if (os.path.isdir ("Gráficos") == 0):
        os.mkdir ("Gráficos")

    i = 1
    k = planilha3.max_row + 3
    l = 0
    while (i <= planilha3.max_column):
        titulo = planilha3.cell (row = 1, column = i).value
        j = 3
        while (j <= planilha3.max_row):
            semestre = planilha3.cell (row = j, column = i).value      
            porcentagem = planilha3.cell (row = j, column = i + 1).value            
            valor = porcentagem*100
            categoriasPlan3.append (semestre)                    
            valoresPlan3.append (valor)            
            j = j + 1

        if (nomeArquivo == "Dados quantitativos agregados (eletivas e clínicas).xlsx"):
            caminho_fig = "Gráficos/(Eletivas) " + titulo + ".png"
            plt.rc ('xtick', labelsize = 6.4)
            fig, ax = plt.subplots (figsize = (6.95, 3.66))
        else:
            caminho_fig = "Gráficos/(Obrigatórias) " + titulo + ".png"
            plt.rc ('xtick', labelsize = 5.95)
            fig, ax = plt.subplots (figsize = (7.4, 3.66))               

        ax.bar (categoriasPlan3, valoresPlan3, width = 0.50)
        ax.set_ylabel ("Porcentagem da pontuação máxima", fontsize = 9.16)
        ax.set_ylim ([0, 100])
        fig.suptitle (titulo)        
        plt.close ()
 
        fig.savefig (caminho_fig)             
        img = openpyxl.drawing.image.Image (caminho_fig)
        if (i % 2 == 1):
            string = str (k)
            string = "A" + string                              
            planilha3.add_image (img, string)            
        else:
            string = str (k)
            string = "M" + string                             
            planilha3.add_image (img, string)
            k = k + 21

        valoresPlan3 = []
        categoriasPlan3 = []
        l = l + 1       
        i = i + 3


def main ():
    criadosObrigatorias = 0
    criadosEletivas = 0
    if (os.path.isfile ("Dados quantitativos agregados (obrigatórias).xlsx")):
        arquivo = openpyxl.load_workbook ("Dados quantitativos agregados (obrigatórias).xlsx")
        nomeArquivo = "Dados quantitativos agregados (obrigatórias).xlsx"       

        criaGraficosSemestres (arquivo, nomeArquivo)
        criaGraficosQuesitos (arquivo, nomeArquivo)

        arquivo.save ("Dados quantitativos agregados (obrigatórias).xlsx")    
        arquivo.close ()
        criadosObrigatorias = 1
    else:
        print ("O arquivo de dados quantitativos agregados das obrigatórias não está disponível!")
        time.sleep (2.5)
    
    if (os.path.isfile ("Dados quantitativos agregados (eletivas e clínicas).xlsx")):
        arquivo = openpyxl.load_workbook ("Dados quantitativos agregados (eletivas e clínicas).xlsx")
        nomeArquivo = "Dados quantitativos agregados (eletivas e clínicas).xlsx"
      
        criaGraficosSemestres (arquivo, nomeArquivo)
        criaGraficosQuesitos (arquivo, nomeArquivo)

        arquivo.save ("Dados quantitativos agregados (eletivas e clínicas).xlsx")    
        arquivo.close ()
        criadosEletivas = 1
    else:
        print ("O arquivo de dados quantitativos agregados das eletivas e clínicas não está disponível!")
        time.sleep (2.5)

    if (criadosObrigatorias == 1 and criadosEletivas == 0):
        print ("Gráficos das obrigatórias criados!")
        time.sleep (2.5)
    elif (criadosObrigatorias == 0 and criadosEletivas == 1):
        print ("Gráficos das eletivas criados!")
        time.sleep (2.5)
    elif (criadosObrigatorias == 1 and criadosEletivas == 1):
        print ("Gráficos das obrigatórias e das eletivas criados!")
        time.sleep (2.5)

main ()
